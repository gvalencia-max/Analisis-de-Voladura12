"""
app.py  ─  SPC Blasting Analysis — Streamlit Web Application
Universidad Nacional del Altiplano · Facultad de Ingeniería de Minas
Estudiante: Giovany Valencia Huanca · Puno, Perú — 2026

Convierte el script CLI original en una aplicación web interactiva con:
  · Sidebar de parámetros con validación en tiempo real
  · Dashboard técnico con 4 secciones expandibles
  · Todos los gráficos matplotlib renderizados inline
  · Exportación de Excel con un solo clic
  · Diagrama técnico de taladros (portada visual)

Ejecutar:
    pip install -r requirements.txt
    py -m streamlit run app.py          (Windows)
    streamlit run app.py                (Linux / macOS)
"""

# ─────────────────────────────────────────────────────────────────────────────
#  Imports
# ─────────────────────────────────────────────────────────────────────────────
import io
import os
import tempfile
import warnings

import matplotlib
matplotlib.use("Agg")
import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import numpy as np
import streamlit as st
from matplotlib.gridspec import GridSpec
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy import stats

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
#  Streamlit page config  (MUST be first Streamlit call)
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SPC Voladura · UNA Puno",
    page_icon="⛏️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
#  CSS — Diseño industrial/técnico (carbón, acero, naranja minero)
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Rajdhani:wght@400;500;600;700&family=Share+Tech+Mono&family=Barlow:wght@300;400;500;600&display=swap');

/* ── Variables de color ── */
:root {
    --carbon:   #0F1117;
    --steel:    #1A1F2E;
    --panel:    #1E2435;
    --border:   #2D3550;
    --orange:   #F07030;
    --orange2:  #FF8C42;
    --gold:     #D4A017;
    --cyan:     #4FC3F7;
    --green:    #43A047;
    --red:      #E53935;
    --text:     #E8ECF4;
    --muted:    #8A93A8;
    --mono:     'Share Tech Mono', monospace;
    --display:  'Rajdhani', sans-serif;
    --body:     'Barlow', sans-serif;
}

/* ── Base ── */
html, body, .stApp {
    background-color: var(--carbon) !important;
    color: var(--text) !important;
    font-family: var(--body) !important;
}

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: var(--steel) !important;
    border-right: 2px solid var(--border);
}
[data-testid="stSidebar"] .stMarkdown h1,
[data-testid="stSidebar"] .stMarkdown h2,
[data-testid="stSidebar"] .stMarkdown h3 {
    color: var(--orange) !important;
    font-family: var(--display) !important;
    letter-spacing: 1px;
}
[data-testid="stSidebar"] label {
    color: var(--cyan) !important;
    font-family: var(--mono) !important;
    font-size: 0.78rem !important;
}

/* ── Inputs ── */
input[type="number"], .stNumberInput input, .stTextInput input {
    background: var(--panel) !important;
    color: var(--text) !important;
    border: 1px solid var(--border) !important;
    border-radius: 4px !important;
    font-family: var(--mono) !important;
}
input[type="number"]:focus, .stNumberInput input:focus {
    border-color: var(--orange) !important;
    box-shadow: 0 0 0 2px rgba(240,112,48,0.25) !important;
}

/* ── Slider ── */
.stSlider [data-baseweb="slider"] {
    color: var(--orange) !important;
}
.stSlider [data-testid="stThumbValue"] { color: var(--orange) !important; }

/* ── Botones principales ── */
.stButton > button {
    background: linear-gradient(135deg, var(--orange), var(--orange2)) !important;
    color: white !important;
    font-family: var(--display) !important;
    font-weight: 700 !important;
    font-size: 1.05rem !important;
    letter-spacing: 1.5px !important;
    border: none !important;
    border-radius: 6px !important;
    padding: 0.6rem 2rem !important;
    transition: all 0.2s ease !important;
}
.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 20px rgba(240,112,48,0.45) !important;
}

/* ── Download button ── */
.stDownloadButton > button {
    background: linear-gradient(135deg, #1B5E20, #43A047) !important;
    color: white !important;
    font-family: var(--display) !important;
    font-weight: 700 !important;
    letter-spacing: 1px !important;
    border: none !important;
    border-radius: 6px !important;
}

/* ── Métricas (KPI cards) ── */
[data-testid="stMetric"] {
    background: var(--panel) !important;
    border: 1px solid var(--border) !important;
    border-radius: 8px !important;
    padding: 0.8rem 1rem !important;
}
[data-testid="stMetric"] label {
    color: var(--muted) !important;
    font-family: var(--mono) !important;
    font-size: 0.75rem !important;
}
[data-testid="stMetricValue"] {
    color: var(--orange) !important;
    font-family: var(--display) !important;
    font-weight: 700 !important;
}
[data-testid="stMetricDelta"] { font-family: var(--mono) !important; }

/* ── Expanders ── */
.streamlit-expanderHeader {
    background: var(--panel) !important;
    color: var(--text) !important;
    font-family: var(--display) !important;
    font-weight: 600 !important;
    font-size: 1rem !important;
    border-left: 4px solid var(--orange) !important;
    border-radius: 4px !important;
    letter-spacing: 0.5px;
}

/* ── Tablas ── */
.stDataFrame, [data-testid="stDataFrame"] {
    border: 1px solid var(--border) !important;
    border-radius: 6px !important;
}

/* ── Divisores / Headers ── */
h1, h2, h3 {
    font-family: var(--display) !important;
    color: var(--text) !important;
}
h1 { letter-spacing: 2px; }
h2 { color: var(--orange) !important; letter-spacing: 1px; }
h3 { color: var(--cyan) !important; }

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    background: var(--steel) !important;
    border-bottom: 2px solid var(--border) !important;
}
.stTabs [data-baseweb="tab"] {
    font-family: var(--display) !important;
    font-weight: 600 !important;
    color: var(--muted) !important;
    letter-spacing: 0.5px;
}
.stTabs [aria-selected="true"] {
    color: var(--orange) !important;
    border-bottom: 3px solid var(--orange) !important;
}

/* ── Alertas ── */
.stSuccess { background: rgba(67,160,71,0.15) !important; border-left: 4px solid var(--green) !important; }
.stWarning { background: rgba(255,160,0,0.15) !important; border-left: 4px solid var(--gold) !important; }
.stError   { background: rgba(229,57,53,0.15)  !important; border-left: 4px solid var(--red) !important; }
.stInfo    { background: rgba(79,195,247,0.12) !important; border-left: 4px solid var(--cyan) !important; }

/* ── Spinner ── */
.stSpinner > div { border-top-color: var(--orange) !important; }

/* ── Barra de progreso ── */
.stProgress > div > div { background: var(--orange) !important; }

/* ── Scrollbar ── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: var(--steel); }
::-webkit-scrollbar-thumb { background: var(--border); border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: var(--orange); }

/* ── Plots (fondo transparente) ── */
[data-testid="stImage"] img { border-radius: 8px; border: 1px solid var(--border); }

/* ── Área de texto (data input) ── */
textarea {
    background: var(--panel) !important;
    color: var(--text) !important;
    border: 1px solid var(--border) !important;
    font-family: var(--mono) !important;
    font-size: 0.82rem !important;
}

/* ── Selectbox ── */
.stSelectbox select, [data-baseweb="select"] {
    background: var(--panel) !important;
    color: var(--text) !important;
}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
#  Constantes SPC
# ─────────────────────────────────────────────────────────────────────────────
SPC_CONSTANTS = {
    2:  {"A2": 1.880, "D3": 0.000, "D4": 3.267, "d2": 1.128},
    3:  {"A2": 1.023, "D3": 0.000, "D4": 2.575, "d2": 1.693},
    4:  {"A2": 0.729, "D3": 0.000, "D4": 2.282, "d2": 2.059},
    5:  {"A2": 0.577, "D3": 0.000, "D4": 2.114, "d2": 2.326},
    6:  {"A2": 0.483, "D3": 0.000, "D4": 2.004, "d2": 2.534},
    7:  {"A2": 0.419, "D3": 0.076, "D4": 1.924, "d2": 2.704},
    8:  {"A2": 0.373, "D3": 0.136, "D4": 1.864, "d2": 2.847},
    10: {"A2": 0.308, "D3": 0.223, "D4": 1.777, "d2": 3.078},
}

# Estilos matplotlib globales
plt.rcParams.update({
    "figure.dpi": 130,
    "font.family": "DejaVu Sans",
    "axes.titlesize": 11,
    "axes.labelsize": 10,
    "xtick.labelsize": 9,
    "ytick.labelsize": 9,
    "legend.fontsize": 9,
    "axes.grid": True,
    "grid.alpha": 0.25,
    "axes.spines.top": False,
    "axes.spines.right": False,
    "figure.facecolor": "#1E2435",
    "axes.facecolor": "#252B3B",
    "axes.edgecolor": "#2D3550",
    "axes.labelcolor": "#E8ECF4",
    "xtick.color": "#8A93A8",
    "ytick.color": "#8A93A8",
    "text.color": "#E8ECF4",
    "legend.facecolor": "#1A1F2E",
    "legend.edgecolor": "#2D3550",
    "grid.color": "#2D3550",
})

# ─────────────────────────────────────────────────────────────────────────────
#  Funciones de cálculo (idénticas al script corregido)
# ─────────────────────────────────────────────────────────────────────────────

def calc_kuz_ram(A, K, Q, RWS, d, B, S, H, W, BCL, CCL):
    X50_cm = A * (K ** -0.8) * (Q ** (1.0 / 6.0)) * ((115.0 / RWS) ** (19.0 / 20.0))
    X50 = X50_cm * 10.0
    L = BCL + CCL
    n_u = ((2.2 - 14.0 * B / d) * ((1.0 + S / B) / 2.0) ** 0.5
           * (1.0 - W / B) * (L / H) ** 0.3
           * (BCL / L + CCL / L))
    n_u = max(0.5, min(n_u, 2.5))
    Xc = X50 / (np.log(2) ** (1.0 / n_u))
    return {"X50": X50, "n_uniformity": n_u, "Xc": Xc, "L": L}


def calc_rosin_rammler(x_arr, Xc, n_u):
    return 1.0 - np.exp(-(x_arr / Xc) ** n_u)


def calc_ppv(K_s, alpha, R, Qd):
    SD = R / (Qd ** 0.5)
    PPV = K_s * (SD ** (-alpha))
    return {"PPV": PPV, "SD": SD}


def calc_spc(subgroups, n_sub):
    means  = np.array([sg.mean() for sg in subgroups])
    ranges = np.array([sg.max() - sg.min() for sg in subgroups])
    X_bar_bar = means.mean()
    R_bar     = ranges.mean()
    c  = SPC_CONSTANTS[n_sub]
    A2, D3, D4, d2 = c["A2"], c["D3"], c["D4"], c["d2"]
    UCL_X = X_bar_bar + A2 * R_bar
    LCL_X = X_bar_bar - A2 * R_bar
    UCL_R = D4 * R_bar
    LCL_R = D3 * R_bar
    return {
        "means": means, "ranges": ranges,
        "X_bar_bar": X_bar_bar, "R_bar": R_bar,
        "UCL_X": UCL_X, "LCL_X": LCL_X, "CL_X": X_bar_bar,
        "UCL_R": UCL_R, "LCL_R": LCL_R, "CL_R": R_bar,
        "sigma_hat": R_bar / d2,
        "A2": A2, "D3": D3, "D4": D4, "d2": d2,
    }


def calc_capability(X_bar_bar, sigma_hat, USL, LSL):
    Cp  = (USL - LSL) / (6.0 * sigma_hat)
    Cpu = (USL - X_bar_bar) / (3.0 * sigma_hat)
    Cpl = (X_bar_bar - LSL) / (3.0 * sigma_hat)
    return {"Cp": Cp, "Cpu": Cpu, "Cpl": Cpl, "Cpk": min(Cpu, Cpl)}


def nelson_rules(means, UCL_X, LCL_X, CL_X, sigma_hat):
    violations = {}
    n = len(means)
    r1 = [i for i in range(n) if means[i] > UCL_X or means[i] < LCL_X]
    if r1:
        violations[1] = r1
    for sign in [1, -1]:
        for i in range(n - 8):
            if all((means[i + j] - CL_X) * sign > 0 for j in range(9)):
                violations.setdefault(2, []).extend(range(i, i + 9))
    if 2 in violations:
        violations[2] = sorted(set(violations[2]))
    for i in range(n - 5):
        seg = means[i:i + 6]
        if all(seg[j] < seg[j + 1] for j in range(5)) or all(seg[j] > seg[j + 1] for j in range(5)):
            violations.setdefault(3, []).extend(range(i, i + 6))
    if 3 in violations:
        violations[3] = sorted(set(violations[3]))
    return violations


def monte_carlo(p, n_mc):
    rng = np.random.default_rng(42)
    COV = {"A": .15, "K": .12, "Q": .10, "RWS": .05,
           "B": .08, "S": .08, "W": .20, "K_s": .25, "alpha": .10, "Qd": .10}

    def tnorm(mean, cov_key, low, high):
        std = mean * COV[cov_key]
        if std == 0:
            return np.full(n_mc, mean)
        a, b = (low - mean) / std, (high - mean) / std
        return stats.truncnorm.rvs(a, b, loc=mean, scale=std, size=n_mc, random_state=rng)

    def lnorm(mean, cov_key):
        cv = COV[cov_key]; sv = np.log(1 + cv ** 2)
        return np.random.lognormal(np.log(mean) - sv / 2, np.sqrt(sv), size=n_mc)

    A_mc    = tnorm(p["A"],     "A",     0.6,  22.0)
    K_mc    = tnorm(p["K"],     "K",     0.05,  2.0)
    Q_mc    = lnorm(p["Q"],     "Q")
    RWS_mc  = tnorm(p["RWS"],   "RWS",  50.0, 150.0)
    B_mc    = tnorm(p["B"],     "B",     0.5,  20.0)
    S_mc    = tnorm(p["S"],     "S",     0.5,  25.0)
    W_mc    = tnorm(p["W"],     "W",     0.001, 0.2)
    K_s_mc  = lnorm(p["K_s"],   "K_s")
    alp_mc  = tnorm(p["alpha"], "alpha", 0.8,   2.5)
    Qd_mc   = tnorm(p["Qd"],    "Qd",    5.0, 1000.0)
    d_mc    = np.full(n_mc, p["d"])
    H_mc    = np.full(n_mc, p["H"])
    BCL_mc  = np.full(n_mc, p["BCL"])
    CCL_mc  = np.full(n_mc, p["CCL"])
    R_mc    = np.full(n_mc, p["R"])
    L_mc    = BCL_mc + CCL_mc

    X50_mc = (A_mc * (K_mc ** -0.8) * (Q_mc ** (1/6)) *
              ((115.0 / RWS_mc) ** (19/20)) * 10.0)
    n_u_mc = ((2.2 - 14.0 * B_mc / d_mc)
              * ((1.0 + S_mc / B_mc) / 2.0) ** 0.5
              * (1.0 - W_mc / B_mc)
              * (L_mc / H_mc) ** 0.3
              * (BCL_mc / L_mc + CCL_mc / L_mc))
    n_u_mc = np.clip(n_u_mc, 0.5, 2.5)
    SD_mc   = R_mc / (Qd_mc ** 0.5)
    PPV_mc  = K_s_mc * (SD_mc ** (-alp_mc))

    USL = p["X50_target"] + p["X50_tol"]
    LSL = p["X50_target"] - p["X50_tol"]
    pct_X50_ok = np.mean((X50_mc >= LSL) & (X50_mc <= USL)) * 100.0
    pct_PPV_ok = np.mean(PPV_mc <= p["PPV_max"]) * 100.0

    def sdict(arr):
        return {"mean": arr.mean(), "std": arr.std(),
                "P5": np.percentile(arr, 5), "P50": np.percentile(arr, 50),
                "P90": np.percentile(arr, 90), "P95": np.percentile(arr, 95)}

    inputs = {"A": A_mc, "K": K_mc, "Q": Q_mc, "RWS": RWS_mc,
              "B": B_mc, "S": S_mc, "W": W_mc,
              "K_s": K_s_mc, "alpha": alp_mc, "Qd": Qd_mc}
    n_s = min(500, n_mc)
    return {
        "X50": sdict(X50_mc), "PPV": sdict(PPV_mc), "n_u": sdict(n_u_mc),
        "pct_X50_ok": pct_X50_ok, "pct_PPV_ok": pct_PPV_ok,
        "corr_X50": {k: np.corrcoef(v, X50_mc)[0, 1] for k, v in inputs.items()},
        "corr_PPV": {k: np.corrcoef(v, PPV_mc)[0, 1] for k, v in inputs.items()},
        "sample": {k: v[:n_s] for k, v in inputs.items()} | {
            "X50": X50_mc[:n_s], "PPV": PPV_mc[:n_s],
            "n_uniformity": n_u_mc[:n_s]},
        "X50_mc": X50_mc, "PPV_mc": PPV_mc,
    }


# ─────────────────────────────────────────────────────────────────────────────
#  Funciones de graficado (devuelven Figure para st.pyplot)
# ─────────────────────────────────────────────────────────────────────────────

def _fig_blast_diagram(p, kuz):
    """Diagrama técnico de taladros — portada visual."""
    B = p["B"]; S = p["S"]; H = p["H"]
    BCL = p["BCL"]; CCL = p["CCL"]; d_mm = p["d"]
    d_m = d_mm / 1000.0
    stemming = max(H - BCL - CCL, 0.0)
    n_col = 4; n_row = 3
    C_ROCK = "#8B7355"; C_BENCH = "#5D4E37"
    C_BCL = "#F07030";  C_CCL  = "#FFB347"
    C_STEM = "#607D8B"; C_FREE = "#4FC3F7"
    C_HOLE = "#1A1F2E"; C_TITLE = "#E8ECF4"

    fig = plt.figure(figsize=(15, 7))
    fig.suptitle(
        "DIAGRAMA TÉCNICO DE VOLADURA EN BANCO  ·  Kuz-Ram / Sadovsky / SPC",
        fontsize=13, fontweight="bold", color=C_TITLE, y=0.97,
        fontfamily="DejaVu Sans"
    )
    gs = GridSpec(2, 3, figure=fig,
                  left=0.04, right=0.97, top=0.91, bottom=0.06,
                  wspace=0.32, hspace=0.40)

    # ── Vista frontal ────────────────────────────────────────────────────
    ax_f = fig.add_subplot(gs[:, 0])
    ax_f.set_facecolor("#1A2535")
    w_total = (n_col - 1) * S + B + 1.5
    ax_f.set_xlim(-0.5, w_total)
    ax_f.set_ylim(-H * 0.14, H * 1.22)
    ax_f.set_aspect("equal")
    ax_f.set_title("Vista Frontal — Sección Transversal", fontsize=9,
                    fontweight="bold", color="#F07030", pad=5)
    ax_f.set_xlabel("Distancia horizontal (m)", fontsize=8, color="#8A93A8")
    ax_f.set_ylabel("Altura (m)", fontsize=8, color="#8A93A8")

    ax_f.add_patch(plt.Rectangle((-0.5, H), w_total + 0.5, H * 0.22,
                                   color=C_BENCH, zorder=1))
    ax_f.add_patch(plt.Rectangle((-0.5, 0), w_total + 0.5, H,
                                   color=C_ROCK, zorder=1, alpha=0.8))
    ax_f.add_patch(plt.Rectangle((-0.5, -H * 0.14), w_total + 0.5, H * 0.14,
                                   color="#3E2723", zorder=1, alpha=0.7))
    ax_f.add_patch(plt.Rectangle(
        ((n_col - 1) * S + B, -H * 0.14), 0.22, H * 1.36,
        color=C_FREE, alpha=0.45, zorder=2, label="Cara libre"))

    for i in range(n_col):
        x0 = i * S
        if stemming > 0:
            ax_f.add_patch(plt.Rectangle(
                (x0 - d_m / 2, H - stemming), d_m, stemming,
                color=C_STEM, zorder=3, label="Taco" if i == 0 else ""))
        ax_f.add_patch(plt.Rectangle(
            (x0 - d_m / 2, BCL), d_m, CCL,
            color=C_CCL, zorder=3, label="CCL" if i == 0 else ""))
        ax_f.add_patch(plt.Rectangle(
            (x0 - d_m / 2, 0), d_m, BCL,
            color=C_BCL, zorder=3, label="BCL" if i == 0 else ""))
        ax_f.add_patch(plt.Rectangle(
            (x0 - d_m / 2, 0), d_m, H,
            edgecolor="#F07030", facecolor="none", lw=0.9, zorder=4))
        ax_f.text(x0, H + H * 0.04, f"T{i+1}", ha="center", fontsize=7,
                   color="#F07030", fontweight="bold")

    # Cotas
    xl = (n_col - 1) * S; xf = xl + B; yc = H * 1.12
    ax_f.annotate("", xy=(xf, yc), xytext=(xl, yc),
                   arrowprops=dict(arrowstyle="<->", color="#4FC3F7", lw=1.5))
    ax_f.text((xl + xf) / 2, yc + H * 0.02, f"B = {B:.1f} m",
               ha="center", fontsize=8, color="#4FC3F7", fontweight="bold")
    ax_f.annotate("", xy=(-0.38, H), xytext=(-0.38, 0),
                   arrowprops=dict(arrowstyle="<->", color="#E8ECF4", lw=1.3))
    ax_f.text(-0.48, H / 2, f"H = {H:.1f} m", ha="right", fontsize=8,
               color="#E8ECF4", fontweight="bold", rotation=90, va="center")
    xi = (n_col // 2) * S
    ax_f.annotate("", xy=(xi + d_m * 2, BCL),
                   xytext=(xi + d_m * 2, 0),
                   arrowprops=dict(arrowstyle="<->", color=C_BCL, lw=1.1))
    ax_f.text(xi + d_m * 3, BCL / 2, f"BCL\n{BCL:.1f} m",
               fontsize=6.5, color=C_BCL, va="center")
    ax_f.annotate("", xy=(xi + d_m * 2, BCL + CCL),
                   xytext=(xi + d_m * 2, BCL),
                   arrowprops=dict(arrowstyle="<->", color=C_CCL, lw=1.1))
    ax_f.text(xi + d_m * 3, BCL + CCL / 2, f"CCL\n{CCL:.1f} m",
               fontsize=6.5, color="#CC6B00", va="center")
    ax_f.legend(loc="lower left", fontsize=7, framealpha=0.7)
    ax_f.tick_params(labelsize=7, colors="#8A93A8")

    # ── Vista de planta ───────────────────────────────────────────────────
    ax_p = fig.add_subplot(gs[0, 1])
    ax_p.set_facecolor("#1A2535")
    mg = 0.4
    ax_p.set_xlim(-mg, (n_col - 1) * S + mg)
    ax_p.set_ylim(-mg, (n_row - 1) * B + B + mg)
    ax_p.set_aspect("equal")
    ax_p.set_title("Vista de Planta — Malla de Perforación", fontsize=9,
                    fontweight="bold", color="#F07030", pad=4)
    ax_p.set_xlabel("Espaciamiento S (m)", fontsize=8, color="#8A93A8")
    ax_p.set_ylabel("Burden B (m)", fontsize=8, color="#8A93A8")

    for row in range(n_row):
        off = (S / 2) * (row % 2)
        for col in range(n_col):
            xh, yh = col * S + off, row * B
            ax_p.add_patch(plt.Circle((xh, yh), d_m * 9, color="#F07030", zorder=3))
            ax_p.add_patch(plt.Circle((xh, yh), d_m * 10, color="#F07030",
                                        fill=False, lw=0.6, zorder=3, alpha=0.4))
        if row > 0:
            for col in range(n_col):
                o1 = (S / 2) * (row % 2); o0 = (S / 2) * ((row - 1) % 2)
                ax_p.plot([col * S + o0, col * S + o1],
                           [(row - 1) * B, row * B], ":", color="#2D3550", lw=0.7)

    ax_p.annotate("", xy=(S, -mg * 0.5), xytext=(0, -mg * 0.5),
                   arrowprops=dict(arrowstyle="<->", color="#4FC3F7", lw=1.2))
    ax_p.text(S / 2, -mg * 0.8, f"S = {S:.1f} m", ha="center",
               fontsize=7, color="#4FC3F7")
    ax_p.tick_params(labelsize=7, colors="#8A93A8")

    # ── Tabla de parámetros ───────────────────────────────────────────────
    ax_t = fig.add_subplot(gs[1, 1])
    ax_t.axis("off")
    ax_t.set_facecolor("#1E2435")
    ax_t.set_title("Parámetros de Diseño", fontsize=9,
                    fontweight="bold", color="#F07030", pad=4)
    rows = [
        ["Factor roca A",    f"{p['A']:.1f}",    "—"],
        ["Powder factor K",  f"{p['K']:.2f}",    "kg/m³"],
        ["Carga Q",          f"{p['Q']:.0f}",    "kg"],
        ["RWS",              f"{p['RWS']:.0f}",  "—"],
        ["Diámetro d",       f"{d_mm:.0f}",      "mm"],
        ["Burden B",         f"{B:.1f}",         "m"],
        ["Espaciamiento S",  f"{S:.1f}",         "m"],
        ["Altura banco H",   f"{H:.1f}",         "m"],
        ["BCL",              f"{BCL:.1f}",        "m"],
        ["CCL",              f"{CCL:.1f}",        "m"],
        ["Taco (calc.)",     f"{stemming:.1f}",   "m"],
        ["X50 predicho",     f"{kuz['X50']:.0f}", "mm"],
        ["Índice n",         f"{kuz['n_uniformity']:.3f}", "—"],
    ]
    tbl = ax_t.table(cellText=rows, colLabels=["Parámetro", "Valor", "Unidad"],
                      loc="center", cellLoc="center")
    tbl.auto_set_font_size(False); tbl.set_fontsize(7.5); tbl.scale(1, 1.15)
    for (ri, ci), cell in tbl.get_celld().items():
        if ri == 0:
            cell.set_facecolor("#1F4E79"); cell.set_text_props(color="white", fontweight="bold")
        elif ri % 2 == 0:
            cell.set_facecolor("#252B3B"); cell.set_text_props(color="#E8ECF4")
        else:
            cell.set_facecolor("#1E2435"); cell.set_text_props(color="#E8ECF4")
        cell.set_edgecolor("#2D3550")

    # ── Panel semáforo ────────────────────────────────────────────────────
    ax_e = fig.add_subplot(gs[:, 2])
    ax_e.axis("off")
    ax_e.set_facecolor("#1E2435")
    ax_e.set_title("Estado del Proceso — Resumen Ejecutivo",
                    fontsize=9, fontweight="bold", color="#F07030", pad=5)

    USL = p["X50_target"] + p["X50_tol"]
    LSL = p["X50_target"] - p["X50_tol"]
    checks = [
        ("X50 predicho (Kuz-Ram)", f"{kuz['X50']:.0f} mm",
         f"Esp: {LSL:.0f}–{USL:.0f} mm", LSL <= kuz["X50"] <= USL),
        ("Índice uniformidad n",   f"{kuz['n_uniformity']:.3f}",
         "Óptimo: 0.8–1.6", 0.8 <= kuz["n_uniformity"] <= 1.6),
        ("Diámetro taladro d",     f"{d_mm:.0f} mm",
         "Rango: 76–311 mm", 76 <= d_mm <= 311),
        ("Relación S/B",           f"{S/B:.2f}",
         "Típico: 1.0–1.3", 1.0 <= S / B <= 1.4),
        ("Aprovechamiento L/H",    f"{(BCL+CCL)/H:.2f}",
         "Óptimo ≥ 0.70", (BCL + CCL) / H >= 0.70),
        ("Taco mínimo",            f"{stemming:.1f} m",
         "Mínimo 20·d = " + f"{20*d_m:.2f} m", stemming >= 20 * d_m),
    ]
    y0 = 0.95; dy = 0.145
    for label, val, ref, ok in checks:
        col = "#43A047" if ok else "#E53935"
        ico = "✔" if ok else "✘"
        ax_e.add_patch(mpatches.FancyBboxPatch(
            (0.02, y0 - dy + 0.01), 0.96, dy - 0.02,
            boxstyle="round,pad=0.01", facecolor=col + "18",
            edgecolor=col, lw=1.4,
            transform=ax_e.transAxes, clip_on=False))
        ax_e.text(0.50, y0 - dy / 2 + dy * 0.30, label, ha="center",
                   fontsize=8, fontweight="bold", color="#E8ECF4",
                   transform=ax_e.transAxes)
        ax_e.text(0.50, y0 - dy / 2, val, ha="center",
                   fontsize=10, fontweight="bold", color=col,
                   transform=ax_e.transAxes)
        ax_e.text(0.50, y0 - dy / 2 - dy * 0.28, ref, ha="center",
                   fontsize=6.5, color="#8A93A8", transform=ax_e.transAxes)
        ax_e.text(0.50, y0 - dy / 2 - dy * 0.46,
                   f"{ico}  {'CUMPLE' if ok else 'REVISAR'}", ha="center",
                   fontsize=7.5, fontweight="bold", color=col,
                   transform=ax_e.transAxes)
        y0 -= dy

    fig.text(0.5, 0.005,
             "Kuz-Ram (Cunningham, 1987)  ·  Sadovsky (1959)  ·  "
             "SPC (Shewhart, 1931)  ·  ISO 22514-2  ·  Montgomery (2013)",
             ha="center", fontsize=6.5, color="#8A93A8", style="italic")
    return fig


def _fig_kuz_ram(kuz):
    x_arr = np.linspace(1, max(kuz["Xc"] * 3, 1000), 500)
    passing = calc_rosin_rammler(x_arr, kuz["Xc"], kuz["n_uniformity"]) * 100
    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.plot(x_arr, passing, color="#4FC3F7", lw=2.2,
            label=f"Rosin-Rammler  n = {kuz['n_uniformity']:.3f}")
    ax.axvline(kuz["X50"], color="#F07030", ls="--", lw=1.8,
               label=f"X₅₀ = {kuz['X50']:.0f} mm")
    ax.axhline(50, color="#8A93A8", ls=":", lw=1)
    ax.set_xlabel("Tamaño de fragmento x (mm)"); ax.set_ylabel("Pasante acumulado (%)")
    ax.set_title("Curva Granulométrica — Modelo Kuz-Ram (Rosin-Rammler)")
    ax.legend(); ax.set_xscale("log"); ax.set_xlim(left=1); ax.set_ylim(0, 100)
    fig.tight_layout(); return fig


def _fig_ppv(p, ppv_res):
    SD_arr = np.linspace(1, 150, 400)
    PPV_arr = p["K_s"] * (SD_arr ** (-p["alpha"]))
    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.loglog(SD_arr, PPV_arr, color="#4FC3F7", lw=2.2, label="Curva Sadovsky")
    ax.axhline(p["PPV_max"], color="#E53935", ls="--", lw=1.8,
               label=f"PPV_máx = {p['PPV_max']} mm/s")
    ax.scatter([ppv_res["SD"]], [ppv_res["PPV"]], color="#F07030",
               s=90, zorder=5, label=f"Punto diseño: SD={ppv_res['SD']:.1f}, PPV={ppv_res['PPV']:.2f}")
    ax.set_xlabel("Distancia escalada SD = R/√Qd  (m/kg⁰·⁵)")
    ax.set_ylabel("PPV (mm/s)")
    ax.set_title("Atenuación de Vibraciones — Modelo de Sadovsky (1959)")
    ax.legend(); fig.tight_layout(); return fig


def _fig_xbar(spc, p):
    means = spc["means"]; rngs = spc["ranges"]
    x = np.arange(1, len(means) + 1)
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 6), sharex=True)
    for ax, data, ucl, lcl, cl, yl, ttl in [
        (ax1, means, spc["UCL_X"], spc["LCL_X"], spc["CL_X"],
         "X̄ (mm)", "Carta X̄ — Medias de Subgrupo"),
        (ax2, rngs, spc["UCL_R"], spc["LCL_R"], spc["CL_R"],
         "R (mm)", "Carta R — Rangos de Subgrupo"),
    ]:
        ax.plot(x, data, color="#4FC3F7", lw=1.5, marker="o", ms=6, label="Datos")
        ax.axhline(ucl, color="#E53935", ls="--", lw=1.5, label=f"UCL={ucl:.1f}")
        ax.axhline(lcl, color="#E53935", ls="--", lw=1.5, label=f"LCL={lcl:.1f}")
        ax.axhline(cl,  color="#43A047", ls="-",  lw=1.5, label=f"CL={cl:.1f}")
        out = [i for i, v in enumerate(data) if v > ucl or v < lcl]
        if out:
            ax.scatter([x[i] for i in out], [data[i] for i in out],
                        color="#F07030", s=110, zorder=6, label="Fuera de control")
        ax.set_ylabel(yl); ax.set_title(ttl)
        ax.legend(loc="upper right", fontsize=8)
    ax2.set_xlabel("Número de subgrupo"); ax2.set_xticks(x)
    fig.tight_layout(); return fig


def _fig_capability(spc, p, cap):
    USL = p["X50_target"] + p["X50_tol"]
    LSL = p["X50_target"] - p["X50_tol"]
    mu = spc["X_bar_bar"]; sigma = spc["sigma_hat"]
    x = np.linspace(mu - 5 * sigma, mu + 5 * sigma, 500)
    y = stats.norm.pdf(x, mu, sigma)
    fig, ax = plt.subplots(figsize=(9, 4.5))
    ax.plot(x, y, color="#4FC3F7", lw=2.2, label=f"Proceso N({mu:.1f}, {sigma:.1f}²)")
    ax.fill_between(x, y, where=(x >= LSL) & (x <= USL), alpha=0.2, color="#43A047", label="Dentro espec.")
    ax.fill_between(x, y, where=(x < LSL) | (x > USL), alpha=0.3, color="#E53935", label="Fuera espec.")
    ax.axvline(USL, color="#E53935", ls="--", lw=1.5, label=f"USL={USL:.0f}")
    ax.axvline(LSL, color="#E53935", ls="--", lw=1.5, label=f"LSL={LSL:.0f}")
    ax.axvline(mu,  color="#4FC3F7", ls="-",  lw=1.2, label=f"X̄={mu:.1f}")
    ax.axvline(p["X50_target"], color="#D4A017", ls=":", lw=1.2,
               label=f"Nominal={p['X50_target']:.0f}")
    ax.set_xlabel("X50 (mm)"); ax.set_ylabel("Densidad de probabilidad")
    ax.set_title(f"Capacidad del Proceso — Cp={cap['Cp']:.2f}  Cpk={cap['Cpk']:.2f}")
    ax.legend(fontsize=8); fig.tight_layout(); return fig


def _fig_mc_x50(mc, p):
    USL = p["X50_target"] + p["X50_tol"]
    LSL = p["X50_target"] - p["X50_tol"]
    arr = mc["X50_mc"]
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4.5))
    cnt, bins, patches = ax1.hist(arr, bins=60, density=True,
                                    color="#4FC3F7", alpha=0.7, edgecolor="none")
    for p2, lft in zip(patches, bins[:-1]):
        if lft < LSL or lft > USL:
            p2.set_facecolor("#E53935")
    ax1.axvline(USL, color="#E53935", ls="--", lw=1.5, label=f"USL={USL:.0f}")
    ax1.axvline(LSL, color="#E53935", ls="--", lw=1.5, label=f"LSL={LSL:.0f}")
    ax1.axvline(mc["X50"]["mean"], color="#4FC3F7", ls="-", lw=1.5,
                label=f"Media={mc['X50']['mean']:.0f}")
    ax1.set_xlabel("X50 (mm)"); ax1.set_ylabel("Densidad")
    ax1.set_title("Monte Carlo — Histograma X50"); ax1.legend(fontsize=8)

    sorted_x = np.sort(arr)
    cdf = np.arange(1, len(sorted_x) + 1) / len(sorted_x)
    ax2.plot(sorted_x, cdf * 100, color="#4FC3F7", lw=1.8)
    ax2.axvline(USL, color="#E53935", ls="--", lw=1.5, label=f"USL={USL:.0f}")
    ax2.axvline(LSL, color="#E53935", ls="--", lw=1.5, label=f"LSL={LSL:.0f}")
    for pct, col in [(5, "#D4A017"), (50, "#43A047"), (95, "#AB47BC")]:
        v = mc["X50"][f"P{pct}"]
        ax2.axvline(v, color=col, ls=":", lw=1.3, label=f"P{pct}={v:.0f}")
    ax2.set_xlabel("X50 (mm)"); ax2.set_ylabel("Probabilidad acumulada (%)")
    ax2.set_title("Monte Carlo — FDA X50"); ax2.legend(fontsize=8)
    fig.tight_layout(); return fig


def _fig_mc_ppv(mc, p):
    arr = mc["PPV_mc"]
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4.5))
    cnt, bins, patches = ax1.hist(arr, bins=60, density=True,
                                    color="#F07030", alpha=0.7, edgecolor="none")
    for p2, lft in zip(patches, bins[:-1]):
        if lft > p["PPV_max"]:
            p2.set_facecolor("#E53935")
    ax1.axvline(p["PPV_max"], color="#E53935", ls="--", lw=1.5,
                label=f"PPV_máx={p['PPV_max']}")
    ax1.axvline(mc["PPV"]["mean"], color="#4FC3F7", ls="-", lw=1.5,
                label=f"Media={mc['PPV']['mean']:.2f}")
    ax1.set_xlabel("PPV (mm/s)"); ax1.set_ylabel("Densidad")
    ax1.set_title("Monte Carlo — Histograma PPV"); ax1.legend(fontsize=8)

    sorted_p = np.sort(arr)
    cdf = np.arange(1, len(sorted_p) + 1) / len(sorted_p)
    ax2.plot(sorted_p, cdf * 100, color="#F07030", lw=1.8)
    ax2.axvline(p["PPV_max"], color="#E53935", ls="--", lw=1.5,
                label=f"PPV_máx={p['PPV_max']}")
    for pct, col in [(5, "#43A047"), (50, "#4FC3F7"), (95, "#AB47BC")]:
        v = mc["PPV"][f"P{pct}"]
        ax2.axvline(v, color=col, ls=":", lw=1.3, label=f"P{pct}={v:.2f}")
    ax2.set_xlabel("PPV (mm/s)"); ax2.set_ylabel("Probabilidad acumulada (%)")
    ax2.set_title("Monte Carlo — FDA PPV"); ax2.legend(fontsize=8)
    fig.tight_layout(); return fig


def _fig_tornado(mc, for_ppv=False):
    key = "corr_PPV" if for_ppv else "corr_X50"
    lbl_map = {"A": "Factor roca A", "K": "Powder factor K",
               "Q": "Carga Q", "RWS": "RWS", "B": "Burden B",
               "S": "Espaciamiento S", "W": "Desv. perforación W",
               "K_s": "Coef. sitio K_s", "alpha": "Atenuación α",
               "Qd": "Carga/retardo Qd"}
    items = sorted(mc[key].items(), key=lambda x: abs(x[1]), reverse=True)
    names = [lbl_map.get(k, k) for k, _ in items]
    vals  = [v for _, v in items]
    fig, ax = plt.subplots(figsize=(9, 5.5))
    base = "#F07030" if for_ppv else "#4FC3F7"
    colors = [base if v >= 0 else "#E53935" for v in vals]
    ax.barh(np.arange(len(names)), vals, color=colors,
             edgecolor="none", height=0.6)
    ax.set_yticks(np.arange(len(names))); ax.set_yticklabels(names)
    ax.axvline(0, color="#E8ECF4", lw=0.8)
    ax.set_xlabel(f"Correlación Pearson con {'PPV' if for_ppv else 'X50'}")
    ax.set_title(f"Análisis de Sensibilidad — {'PPV' if for_ppv else 'X50'} (Tornado Monte Carlo)")
    ax.set_xlim(-1.05, 1.05); fig.tight_layout(); return fig


def _fig_scatter(mc):
    sample = mc["sample"]
    top = ["A", "K", "Q", "B"]
    X50_s = sample["X50"]
    lbl = {"A": "Factor roca A (—)", "K": "Powder factor K (kg/m³)",
           "Q": "Carga Q (kg)", "B": "Burden B (m)"}
    fig, axes = plt.subplots(2, 2, figsize=(10, 7))
    axes = axes.flatten()
    for i, var in enumerate(top):
        ax = axes[i]
        ax.scatter(sample[var], X50_s, alpha=0.25, s=7, color="#4FC3F7")
        m, b = np.polyfit(sample[var], X50_s, 1)
        xf = np.linspace(sample[var].min(), sample[var].max(), 100)
        ax.plot(xf, m * xf + b, color="#F07030", lw=1.5,
                label=f"r = {mc['corr_X50'][var]:.2f}")
        ax.set_xlabel(lbl[var]); ax.set_ylabel("X50 (mm)")
        ax.legend(fontsize=9)
    fig.suptitle("Diagramas de Dispersión — Inputs vs X50 (Monte Carlo)", fontsize=11)
    fig.tight_layout(); return fig


def _fig_capability_bars(cap):
    fig, ax = plt.subplots(figsize=(6, 4))
    idxs = ["Cp", "Cpu", "Cpl", "Cpk"]
    vals  = [cap["Cp"], cap["Cpu"], cap["Cpl"], cap["Cpk"]]
    colors = ["#43A047" if v >= 1.33 else "#FFB347" if v >= 1.0 else "#E53935"
               for v in vals]
    bars = ax.bar(idxs, vals, color=colors, edgecolor="none", width=0.5)
    ax.axhline(1.33, color="#E8ECF4", ls="--", lw=1.5, label="Umbral 1.33")
    ax.axhline(1.00, color="#8A93A8", ls=":", lw=1.2, label="Mínimo 1.00")
    ax.set_ylabel("Índice de capacidad")
    ax.set_title("Índices de Capacidad del Proceso")
    ax.legend(fontsize=9)
    for bar, val in zip(bars, vals):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.02,
                f"{val:.2f}", ha="center", va="bottom",
                fontsize=10, fontweight="bold", color="#E8ECF4")
    fig.tight_layout(); return fig


# ─────────────────────────────────────────────────────────────────────────────
#  Exportación Excel (en memoria → bytes para st.download_button)
# ─────────────────────────────────────────────────────────────────────────────
def _build_excel(p, kuz, ppv_res, spc, cap, mc):
    HFILL = PatternFill("solid", fgColor="1F4E79")
    HFONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    SFILL = PatternFill("solid", fgColor="BDD7EE")
    SFONT = Font(bold=True, name="Calibri", size=10)
    GFILL = PatternFill("solid", fgColor="C6EFCE")
    RFILL = PatternFill("solid", fgColor="FFC7CE")
    YFILL = PatternFill("solid", fgColor="FFEB9C")
    WFILL = PatternFill("solid", fgColor="FFFFFF")
    LFILL = PatternFill("solid", fgColor="F2F2F2")

    def hrow(ws, row, vals):
        for j, v in enumerate(vals):
            c = ws.cell(row=row, column=1 + j, value=v)
            c.fill = HFILL; c.font = HFONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def srow(ws, row, lbl, nc=5):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
        c = ws.cell(row=row, column=1, value=lbl)
        c.fill = SFILL; c.font = SFONT
        c.alignment = Alignment(horizontal="left", vertical="center")

    def drow(ws, row, vals, fill=None):
        for j, v in enumerate(vals):
            c = ws.cell(row=row, column=1 + j, value=v)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if fill: c.fill = fill

    def aw(ws):
        for col in ws.columns:
            ml = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, 55)

    USL = p["X50_target"] + p["X50_tol"]
    LSL = p["X50_target"] - p["X50_tol"]
    wb = Workbook(); wb.remove(wb.active)

    # Hoja 1 — Datos
    ws = wb.create_sheet("1_Datos_Entrada")
    ws.sheet_view.showGridLines = False
    hrow(ws, 1, ["Parámetro / Símbolo", "Valor", "Unidades", "Valores típicos", "CoV MC (%)"])
    rows_in = [
        ("─── KUZ-RAM ───", []),
        ("Factor de roca A", p["A"], "—", "Blanda:0.6–3 | Media:3–6 | Dura:6–14 | Muy dura:14–22", "15%"),
        ("Powder factor K",  p["K"], "kg/m³", "0.10–0.25 / 0.25–0.50 / 0.50–0.80", "12%"),
        ("Carga Q",          p["Q"], "kg",    "50–100 / 100–200 / 200–500", "10%"),
        ("RWS",              p["RWS"], "—",   "ANFO:100 | HANFO:110–120 | Emulsión:115–125", "5%"),
        ("Diámetro d",       p["d"],  "mm",   "76–102 / 115–165 / 200–311", "0%"),
        ("Burden B",         p["B"],  "m",    "2–3 / 3–5 / 5–8", "8%"),
        ("Espaciamiento S",  p["S"],  "m",    "S≈B / 1.15B / 1.25–1.5B", "8%"),
        ("Altura banco H",   p["H"],  "m",    "5–8 / 8–15 / 15–20", "0%"),
        ("Desv. perforación W", p["W"], "m",  "0.01–0.02 / 0.02–0.04 / 0.04–0.06", "20%"),
        ("BCL",              p["BCL"], "m",   "1–2 / 2–3.5 / 3–5", "0%"),
        ("CCL",              p["CCL"], "m",   "2–5 / 5–10 / 8–15", "0%"),
        ("─── SADOVSKY ───", []),
        ("K_s (coef. sitio)", p["K_s"], "—",  "Roca dura:100–300 | Media:200–500 | Fracturada:400–800", "25%"),
        ("α (atenuación)",   p["alpha"], "—", "Dura:1.7–2.0 | Media:1.5–1.8 | Fracturada:1.2–1.5", "10%"),
        ("Distancia R",      p["R"],   "m",   "50–100 / 100–500 / 500–2000", "0%"),
        ("Carga/retardo Qd", p["Qd"],  "kg",  "20–60 / 60–150 / 150–300", "10%"),
        ("PPV_máx",          p["PPV_max"], "mm/s", "NTP:15 | USBM:19–50.8 | DIN:50", "0%"),
        ("─── SPC ───", []),
        ("X50 nominal",      p["X50_target"], "mm", "80–150 / 150–350 / 300–600", "0%"),
        ("Tolerancia ±tol",  p["X50_tol"],    "mm", "±30–60 / ±60–100 / ±100–150", "0%"),
        ("USL (calculado)",  USL, "mm", "Automático", "—"),
        ("LSL (calculado)",  LSL, "mm", "Automático", "—"),
        ("Tamaño subgrupo n", p["n_subgroup"], "voladuras", str(list(SPC_CONSTANTS.keys())), "0%"),
    ]
    r = 2; fill = LFILL
    for row in rows_in:
        if isinstance(row[1], list):
            srow(ws, r, row[0]); r += 1; continue
        drow(ws, r, list(row), fill=fill)
        fill = WFILL if fill == LFILL else LFILL; r += 1
    aw(ws)

    # Hoja 2 — Resultados
    ws2 = wb.create_sheet("2_Resultados")
    ws2.sheet_view.showGridLines = False
    hrow(ws2, 1, ["Análisis", "Variable", "Valor", "Interpretación", "Recomendación"])
    det = [
        ("Kuz-Ram", "X50 predicho", f"{kuz['X50']:.0f} mm",
         f"X50={kuz['X50']:.0f} mm. Espec.: {LSL:.0f}–{USL:.0f} mm.",
         "Ajustar K o B/S si X50 fuera de especificación.",
         "green" if LSL <= kuz["X50"] <= USL else "red"),
        ("Kuz-Ram", "Uniformidad n", f"{kuz['n_uniformity']:.3f}",
         f"n={kuz['n_uniformity']:.3f}. {'Buena uniformidad.' if kuz['n_uniformity']>1.0 else 'PSD amplia.'}",
         "Mejorar relación S/B para aumentar n.",
         "green" if kuz["n_uniformity"] > 0.8 else "yellow"),
        ("Kuz-Ram", "Xc", f"{kuz['Xc']:.0f} mm",
         f"Tamaño característico Xc={kuz['Xc']:.0f} mm (63.2% pasante).",
         "Ver curva granulométrica.", "green"),
        ("Sadovsky", "SD", f"{ppv_res['SD']:.1f} m/kg⁰·⁵",
         f"SD=R/√Qd={p['R']:.0f}/√{p['Qd']:.0f}={ppv_res['SD']:.1f}.",
         "Aumentar SD reduciendo Qd si PPV es alto.",
         "green" if ppv_res["SD"] > 20 else "yellow"),
        ("Sadovsky", "PPV", f"{ppv_res['PPV']:.2f} mm/s",
         f"PPV={ppv_res['PPV']:.2f} mm/s vs PPV_máx={p['PPV_max']:.0f}. FS={p['PPV_max']/ppv_res['PPV']:.1f}×.",
         "Reducir Qd si PPV > PPV_máx.",
         "green" if ppv_res["PPV"] < p["PPV_max"] else "red"),
        ("SPC", "X̄̄", f"{spc['X_bar_bar']:.1f} mm",
         f"Gran media={spc['X_bar_bar']:.1f} mm. Desv. nominal={spc['X_bar_bar']-p['X50_target']:+.1f} mm.",
         "Centrar proceso en nominal.", "green"),
        ("SPC", "σ̂", f"{spc['sigma_hat']:.2f} mm",
         f"σ̂=R̄/d2={spc['R_bar']:.1f}/{spc['d2']}={spc['sigma_hat']:.2f} mm.",
         "Reducir variabilidad mejorando perforación.", "green"),
        ("Capacidad", "Cp", f"{cap['Cp']:.2f}",
         f"Cp={cap['Cp']:.2f}. {'CAPAZ.' if cap['Cp']>=1.33 else 'NO CAPAZ.'}",
         "Objetivo Cp ≥ 1.33.", "green" if cap["Cp"] >= 1.33 else "red"),
        ("Capacidad", "Cpk", f"{cap['Cpk']:.2f}",
         f"Cpk=min(Cpu={cap['Cpu']:.2f}, Cpl={cap['Cpl']:.2f})={cap['Cpk']:.2f}.",
         "Objetivo Cpk ≥ 1.33.", "green" if cap["Cpk"] >= 1.33 else "orange"),
    ]
    CFILL = {"green": GFILL, "red": RFILL, "orange": YFILL, "yellow": YFILL}
    r = 2
    srow(ws2, r, "─── ANÁLISIS DETERMINÍSTICO ───"); r += 1
    for row in det:
        drow(ws2, r, list(row[:5]))
        for j in range(5):
            ws2.cell(row=r, column=1+j).fill = CFILL.get(row[5], WFILL)
        r += 1
    srow(ws2, r, "─── MONTE CARLO ───"); r += 1
    mc_rows = [
        ("MC — X50", "Media ± σ", f"{mc['X50']['mean']:.0f} ± {mc['X50']['std']:.0f} mm",
         f"P5={mc['X50']['P5']:.0f} P50={mc['X50']['P50']:.0f} P90={mc['X50']['P90']:.0f} P95={mc['X50']['P95']:.0f} mm.",
         f"{mc['pct_X50_ok']:.1f}% dentro especificación.",
         "green" if mc["pct_X50_ok"] > 90 else "yellow"),
        ("MC — PPV", "Media ± σ", f"{mc['PPV']['mean']:.2f} ± {mc['PPV']['std']:.2f} mm/s",
         f"P95={mc['PPV']['P95']:.2f} vs límite {p['PPV_max']:.0f} mm/s.",
         f"{mc['pct_PPV_ok']:.1f}% bajo el límite.",
         "green" if mc["pct_PPV_ok"] > 95 else "orange"),
    ]
    for row in mc_rows:
        drow(ws2, r, list(row[:5]))
        for j in range(5):
            ws2.cell(row=r, column=1+j).fill = CFILL.get(row[5], WFILL)
        r += 1
    aw(ws2)

    # Hoja 3 — SPC
    ws3 = wb.create_sheet("3_SPC_Calculos")
    ws3.sheet_view.showGridLines = False
    hrow(ws3, 1, ["Subgrupo"] + [f"Vol{i+1}" for i in range(p["n_subgroup"])] +
         ["Media X̄i (mm)", "Rango Ri (mm)", "Estado"])
    for i, sg in enumerate(p["subgroups"]):
        rv = [i + 1] + list(sg) + ([""] * (p["n_subgroup"] - len(sg)))
        rv += [round(spc["means"][i], 2), round(spc["ranges"][i], 2)]
        out = spc["means"][i] > spc["UCL_X"] or spc["means"][i] < spc["LCL_X"]
        rv.append("FUERA ⚠" if out else "En control ✓")
        drow(ws3, 2 + i, rv, fill=RFILL if out else GFILL)
    r2 = 2 + len(p["subgroups"]) + 2
    srow(ws3, r2, "─── LÍMITES DE CONTROL ───", nc=p["n_subgroup"] + 4); r2 += 1
    for row in [
        ["Carta X̄", "UCL_X", f"{spc['UCL_X']:.2f}", "CL_X", f"{spc['CL_X']:.2f}",
         "LCL_X", f"{spc['LCL_X']:.2f}", "A2", str(spc["A2"])],
        ["Carta R",  "UCL_R", f"{spc['UCL_R']:.2f}", "CL_R", f"{spc['CL_R']:.2f}",
         "LCL_R", f"{spc['LCL_R']:.2f}", "D3/D4", f"{spc['D3']}/{spc['D4']}"],
        ["Capacidad","σ̂", f"{spc['sigma_hat']:.4f}", "Cp", f"{cap['Cp']:.4f}",
         "Cpk", f"{cap['Cpk']:.4f}", "d2", str(spc["d2"])],
    ]:
        drow(ws3, r2, row); r2 += 1
    aw(ws3)

    # Hoja 4 — Monte Carlo stats
    ws4 = wb.create_sheet("4_MonteCarlo")
    ws4.sheet_view.showGridLines = False
    hrow(ws4, 1, ["Variable", "Media", "σ", "P5", "P50", "P90", "P95", "Unidad"])
    r2 = 2
    for lbl, st, unit in [("X50", mc["X50"], "mm"),
                            ("PPV", mc["PPV"], "mm/s"),
                            ("n uniformidad", mc["n_u"], "—")]:
        drow(ws4, r2, [lbl, round(st["mean"],3), round(st["std"],3),
                       round(st["P5"],3), round(st["P50"],3),
                       round(st["P90"],3), round(st["P95"],3), unit])
        r2 += 1
    r2 += 1
    srow(ws4, r2, "─── CORRELACIONES PEARSON ───", nc=4); r2 += 1
    hrow(ws4, r2, ["Variable", "Corr. X50", "Corr. PPV", "Impacto X50"])
    r2 += 1
    for var in sorted(mc["corr_X50"], key=lambda k: abs(mc["corr_X50"][k]), reverse=True):
        cx = mc["corr_X50"][var]; cp = mc["corr_PPV"].get(var, 0)
        imp = "ALTO" if abs(cx) > 0.5 else "MEDIO" if abs(cx) > 0.2 else "BAJO"
        fill = RFILL if abs(cx) > 0.5 else YFILL if abs(cx) > 0.2 else GFILL
        drow(ws4, r2, [var, round(cx,4), round(cp,4), imp], fill=fill)
        r2 += 1
    aw(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
#  SIDEBAR — Entrada de parámetros
# ─────────────────────────────────────────────────────────────────────────────
def build_sidebar():
    st.sidebar.markdown("""
## ⛏️ SPC VOLADURA
### Universidad Nacional del Altiplano
**Ing. de Minas · Puno, Perú**
---
""")
    p = {}

    with st.sidebar.expander("🪨  SECCIÓN 1 — Kuz-Ram", expanded=True):
        p["A"]   = st.number_input("Factor de roca A (0.6–22)",
                                    min_value=0.6, max_value=22.0, value=8.0, step=0.5)
        p["K"]   = st.number_input("Powder factor K (kg/m³)",
                                    min_value=0.05, max_value=2.0, value=0.40, step=0.05)
        p["Q"]   = st.number_input("Carga por barreno Q (kg)",
                                    min_value=10.0, max_value=1000.0, value=150.0, step=10.0)
        p["RWS"] = st.number_input("RWS (ANFO=100)",
                                    min_value=50.0, max_value=150.0, value=105.0, step=5.0)
        p["d"]   = st.number_input("Diámetro barreno d (mm)",
                                    min_value=50.0, max_value=400.0, value=127.0, step=1.0)
        p["B"]   = st.number_input("Burden B (m)",
                                    min_value=0.5, max_value=20.0, value=4.5, step=0.1)
        p["S"]   = st.number_input("Espaciamiento S (m)",
                                    min_value=0.5, max_value=25.0, value=5.0, step=0.1)
        p["H"]   = st.number_input("Altura de banco H (m)",
                                    min_value=2.0, max_value=50.0, value=10.0, step=0.5)
        p["W"]   = st.number_input("Desv. estándar perforación W (m)",
                                    min_value=0.001, max_value=0.2, value=0.03, step=0.005,
                                    format="%.3f")
        p["BCL"] = st.number_input("Carga de fondo BCL (m)",
                                    min_value=0.5, max_value=15.0, value=3.0, step=0.5)
        p["CCL"] = st.number_input("Carga columna CCL (m)",
                                    min_value=0.5, max_value=30.0, value=7.0, step=0.5)

    with st.sidebar.expander("📳  SECCIÓN 2 — Sadovsky PPV"):
        p["K_s"]    = st.number_input("Coef. sitio K_s (100–1200)",
                                       min_value=50.0, max_value=1500.0, value=450.0, step=25.0)
        p["alpha"]  = st.number_input("Exponente atenuación α (0.8–2.5)",
                                       min_value=0.8, max_value=2.5, value=1.60, step=0.05)
        p["R"]      = st.number_input("Distancia al punto de control R (m)",
                                       min_value=10.0, max_value=3000.0, value=200.0, step=10.0)
        p["Qd"]     = st.number_input("Carga máxima por retardo Qd (kg)",
                                       min_value=5.0, max_value=1000.0, value=100.0, step=5.0)
        p["PPV_max"] = st.number_input("PPV máximo permisible (mm/s)",
                                        min_value=1.0, max_value=200.0, value=25.0, step=1.0)

    with st.sidebar.expander("📊  SECCIÓN 3 — SPC y Especificaciones"):
        p["X50_target"] = st.number_input("X50 nominal de diseño (mm)",
                                           min_value=30.0, max_value=1000.0,
                                           value=300.0, step=10.0)
        p["X50_tol"]    = st.number_input("Tolerancia ±tol (mm)",
                                           min_value=10.0, max_value=300.0,
                                           value=80.0, step=5.0)
        p["n_subgroup"] = st.selectbox("Tamaño de subgrupo n",
                                        options=list(SPC_CONSTANTS.keys()), index=3)
        p["n_mc"]       = st.select_slider("Simulaciones Monte Carlo",
                                            options=[1000, 5000, 10000, 25000, 50000],
                                            value=10000)

    # Datos históricos X50
    with st.sidebar.expander("📋  Datos históricos X50 (mm)"):
        st.markdown(
            "Ingresa los subgrupos separados por comas, uno por línea. "
            "Cada fila debe tener exactamente **n** valores."
        )
        default_txt = (
            "320,295,340,285,310\n"
            "305,280,325,300,290\n"
            "350,330,360,345,340\n"
            "270,295,280,300,285\n"
            "315,330,305,320,340"
        )
        raw = st.text_area("Subgrupos (una fila = un subgrupo)", value=default_txt,
                           height=160, key="raw_subgroups")

    subgroups = []
    errors = []
    n_sub = int(p["n_subgroup"])
    for li, line in enumerate(raw.strip().splitlines()):
        line = line.strip()
        if not line:
            continue
        try:
            vals = [float(v.strip()) for v in line.split(",")]
            if len(vals) != n_sub:
                errors.append(f"Fila {li+1}: se esperan {n_sub} valores, hay {len(vals)}.")
                continue
            subgroups.append(np.array(vals))
        except ValueError:
            errors.append(f"Fila {li+1}: valor no numérico.")

    if errors:
        for e in errors:
            st.sidebar.error(e)
    if len(subgroups) < 3:
        st.sidebar.warning(f"Se necesitan ≥ 3 subgrupos. Actualmente: {len(subgroups)}.")

    p["subgroups"] = subgroups
    return p


# ─────────────────────────────────────────────────────────────────────────────
#  Helpers de UI
# ─────────────────────────────────────────────────────────────────────────────
def _status_badge(ok: bool) -> str:
    if ok:
        return "🟢 CUMPLE"
    return "🔴 REVISAR"


def _capability_badge(v: float) -> str:
    if v >= 1.33:
        return f"🟢 {v:.2f}"
    if v >= 1.00:
        return f"🟡 {v:.2f}"
    return f"🔴 {v:.2f}"


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    # ── Header principal ──────────────────────────────────────────────────
    st.markdown("""
<div style="
    background: linear-gradient(135deg, #0F1117 0%, #1A1F2E 50%, #0F1117 100%);
    border: 1px solid #2D3550;
    border-left: 6px solid #F07030;
    border-radius: 8px;
    padding: 1.4rem 2rem 1rem 2rem;
    margin-bottom: 1.5rem;">
  <h1 style="
      font-family:'Rajdhani',sans-serif;
      font-size:2.1rem;
      font-weight:700;
      letter-spacing:3px;
      color:#E8ECF4;
      margin:0 0 0.2rem 0;">
    ⛏️  SPC BLASTING ANALYSIS
  </h1>
  <p style="
      font-family:'Share Tech Mono',monospace;
      font-size:0.82rem;
      color:#F07030;
      margin:0 0 0.15rem 0;
      letter-spacing:1.5px;">
    KUZ-RAM · SADOVSKY · CONTROL ESTADÍSTICO DE PROCESOS · MONTE CARLO
  </p>
  <p style="
      font-family:'Barlow',sans-serif;
      font-size:0.78rem;
      color:#8A93A8;
      margin:0;">
    Universidad Nacional del Altiplano &nbsp;·&nbsp; Facultad de Ingeniería de Minas
    &nbsp;·&nbsp; Estudiante: Giovany Valencia Huanca &nbsp;·&nbsp; Puno, Perú — 2026
  </p>
</div>
""", unsafe_allow_html=True)

    # ── Sidebar ───────────────────────────────────────────────────────────
    p = build_sidebar()

    # ── Botón de análisis ─────────────────────────────────────────────────
    col_btn, col_info = st.columns([1, 3])
    with col_btn:
        run_ok = len(p["subgroups"]) >= 3
        run = st.button("▶  EJECUTAR ANÁLISIS", disabled=not run_ok,
                        use_container_width=True)
    with col_info:
        if not run_ok:
            st.info("ℹ️  Ingresa al menos **3 subgrupos** de datos históricos X50 para habilitar el análisis.")
        else:
            st.success(f"✓  {len(p['subgroups'])} subgrupos listos — presiona **EJECUTAR ANÁLISIS**")

    if not run:
        st.markdown("---")
        st.markdown("""
<div style="
    text-align:center;
    padding:3rem 1rem;
    color:#2D3550;
    font-family:'Rajdhani',sans-serif;
    font-size:1.1rem;
    letter-spacing:2px;">
  CONFIGURA LOS PARÁMETROS EN EL PANEL LATERAL Y PRESIONA EJECUTAR
</div>
""", unsafe_allow_html=True)
        return

    # ── Cálculos ──────────────────────────────────────────────────────────
    with st.spinner("⚙️  Calculando modelos Kuz-Ram y Sadovsky..."):
        kuz     = calc_kuz_ram(p["A"], p["K"], p["Q"], p["RWS"], p["d"],
                               p["B"], p["S"], p["H"], p["W"], p["BCL"], p["CCL"])
        ppv_res = calc_ppv(p["K_s"], p["alpha"], p["R"], p["Qd"])

    USL = p["X50_target"] + p["X50_tol"]
    LSL = p["X50_target"] - p["X50_tol"]

    with st.spinner("📊  Calculando SPC y capacidad del proceso..."):
        spc        = calc_spc(p["subgroups"], int(p["n_subgroup"]))
        cap        = calc_capability(spc["X_bar_bar"], spc["sigma_hat"], USL, LSL)
        violations = nelson_rules(spc["means"], spc["UCL_X"], spc["LCL_X"],
                                  spc["CL_X"], spc["sigma_hat"])

    with st.spinner(f"🎲  Ejecutando Monte Carlo ({p['n_mc']:,} simulaciones)..."):
        mc = monte_carlo(p, int(p["n_mc"]))

    # ── Diagrama técnico de taladros (portada) ────────────────────────────
    st.markdown("## 🔩 Diagrama Técnico de Diseño de Voladura")
    fig_diag = _fig_blast_diagram(p, kuz)
    st.pyplot(fig_diag, use_container_width=True)
    plt.close(fig_diag)

    # ── KPIs principales ──────────────────────────────────────────────────
    st.markdown("## 📈 Indicadores Clave del Proceso")
    k1, k2, k3, k4, k5, k6 = st.columns(6)
    x50_ok  = LSL <= kuz["X50"] <= USL
    ppv_ok  = ppv_res["PPV"] < p["PPV_max"]
    cp_ok   = cap["Cp"] >= 1.33
    cpk_ok  = cap["Cpk"] >= 1.33

    k1.metric("X50 predicho",      f"{kuz['X50']:.0f} mm",
              f"Espec: {LSL:.0f}–{USL:.0f} mm")
    k2.metric("PPV predicha",       f"{ppv_res['PPV']:.2f} mm/s",
              f"Límite: {p['PPV_max']:.0f} mm/s")
    k3.metric("X̄ (SPC)",           f"{spc['X_bar_bar']:.1f} mm",
              f"σ̂ = {spc['sigma_hat']:.2f} mm")
    k4.metric("Cp (potencial)",     _capability_badge(cap["Cp"]),
              "≥ 1.33 requerido")
    k5.metric("Cpk (real)",         _capability_badge(cap["Cpk"]),
              "≥ 1.33 requerido")
    k6.metric("MC X50 en espec.",   f"{mc['pct_X50_ok']:.1f} %",
              f"PPV ok: {mc['pct_PPV_ok']:.1f} %")

    # Alertas Nelson
    if violations:
        st.markdown("---")
        for rule, idxs in violations.items():
            msgs = {
                1: "**Regla 1 Nelson:** punto(s) fuera de los límites UCL/LCL",
                2: "**Regla 2 Nelson:** 9 puntos consecutivos al mismo lado de la línea central",
                3: "**Regla 3 Nelson:** 6 puntos en tendencia monótona",
            }
            st.warning(f"⚠️ {msgs.get(rule, f'Regla {rule}')}  —  "
                       f"Subgrupos: {[i+1 for i in idxs]}")
    else:
        st.success("✅  Sin violaciones de las Reglas de Nelson detectadas.")

    # ── Tabs con resultados ───────────────────────────────────────────────
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "🪨  Kuz-Ram & PPV",
        "📊  Cartas SPC",
        "⚙️  Capacidad",
        "🎲  Monte Carlo",
        "📋  Tablas",
    ])

    with tab1:
        st.markdown("### Modelo de Fragmentación Kuz-Ram")
        c1, c2 = st.columns([2, 1])
        with c1:
            fig = _fig_kuz_ram(kuz)
            st.pyplot(fig, use_container_width=True); plt.close(fig)
        with c2:
            st.markdown(f"""
| Parámetro | Valor |
|-----------|-------|
| **X50 predicho** | {kuz['X50']:.1f} mm |
| **Índice n** | {kuz['n_uniformity']:.4f} |
| **Xc** | {kuz['Xc']:.1f} mm |
| **LSL / USL** | {LSL:.0f} / {USL:.0f} mm |
| **Estado** | {_status_badge(x50_ok)} |
""")
        st.markdown("### Modelo de Atenuación PPV — Sadovsky (1959)")
        c3, c4 = st.columns([2, 1])
        with c3:
            fig = _fig_ppv(p, ppv_res)
            st.pyplot(fig, use_container_width=True); plt.close(fig)
        with c4:
            sf = p["PPV_max"] / ppv_res["PPV"]
            st.markdown(f"""
| Parámetro | Valor |
|-----------|-------|
| **SD** | {ppv_res['SD']:.2f} m/kg⁰·⁵ |
| **PPV predicha** | {ppv_res['PPV']:.3f} mm/s |
| **PPV_máx** | {p['PPV_max']:.1f} mm/s |
| **Factor seguridad** | {sf:.2f} × |
| **Estado** | {_status_badge(ppv_ok)} |
""")

    with tab2:
        st.markdown("### Cartas de Control X̄-R")
        fig = _fig_xbar(spc, p)
        st.pyplot(fig, use_container_width=True); plt.close(fig)

        st.markdown("### Límites de Control Calculados")
        c1, c2 = st.columns(2)
        c1.markdown(f"""
**Carta X̄** (A₂ = {spc['A2']})

| Límite | Valor |
|--------|-------|
| UCL_X̄ | **{spc['UCL_X']:.2f} mm** |
| CL  | {spc['CL_X']:.2f} mm |
| LCL_X̄ | **{spc['LCL_X']:.2f} mm** |
""")
        c2.markdown(f"""
**Carta R** (D₃={spc['D3']}, D₄={spc['D4']})

| Límite | Valor |
|--------|-------|
| UCL_R | **{spc['UCL_R']:.2f} mm** |
| R̄ | {spc['CL_R']:.2f} mm |
| LCL_R | **{spc['LCL_R']:.2f} mm** |
""")

    with tab3:
        st.markdown("### Análisis de Capacidad del Proceso (Cp, Cpk)")
        c1, c2 = st.columns([2, 1.2])
        with c1:
            fig = _fig_capability(spc, p, cap)
            st.pyplot(fig, use_container_width=True); plt.close(fig)
        with c2:
            fig = _fig_capability_bars(cap)
            st.pyplot(fig, use_container_width=True); plt.close(fig)

        st.markdown(f"""
**Fórmulas aplicadas** (σ̂ = R̄/d₂ = {spc['R_bar']:.2f}/{spc['d2']} = {spc['sigma_hat']:.4f} mm)

| Índice | Fórmula | Valor | Interpretación |
|--------|---------|-------|---------------|
| **Cp** | (USL−LSL)/(6σ̂) = ({USL:.0f}−{LSL:.0f})/(6×{spc['sigma_hat']:.2f}) | **{cap['Cp']:.4f}** | {"✅ Proceso capaz" if cap['Cp']>=1.33 else "⚠️ Marginalmente capaz" if cap['Cp']>=1.0 else "❌ No capaz"} |
| **Cpu** | (USL−X̄)/(3σ̂) | **{cap['Cpu']:.4f}** | Margen superior |
| **Cpl** | (X̄−LSL)/(3σ̂) | **{cap['Cpl']:.4f}** | Margen inferior |
| **Cpk** | min(Cpu, Cpl) | **{cap['Cpk']:.4f}** | {"✅ Proceso capaz" if cap['Cpk']>=1.33 else "⚠️ Marginalmente capaz" if cap['Cpk']>=1.0 else "❌ No capaz"} |
| **Cp−Cpk** | Descentramiento | **{cap['Cp']-cap['Cpk']:.4f}** | {"Proceso bien centrado ✅" if cap['Cp']-cap['Cpk']<0.1 else "Proceso descentrado ⚠️"} |
""")

    with tab4:
        st.markdown("### Monte Carlo — Distribución X50")
        fig = _fig_mc_x50(mc, p)
        st.pyplot(fig, use_container_width=True); plt.close(fig)

        st.markdown("### Monte Carlo — Distribución PPV")
        fig = _fig_mc_ppv(mc, p)
        st.pyplot(fig, use_container_width=True); plt.close(fig)

        st.markdown("### Análisis de Sensibilidad (Tornado)")
        c1, c2 = st.columns(2)
        with c1:
            fig = _fig_tornado(mc, for_ppv=False)
            st.pyplot(fig, use_container_width=True); plt.close(fig)
        with c2:
            fig = _fig_tornado(mc, for_ppv=True)
            st.pyplot(fig, use_container_width=True); plt.close(fig)

        st.markdown("### Dispersión — Inputs vs X50")
        fig = _fig_scatter(mc)
        st.pyplot(fig, use_container_width=True); plt.close(fig)

    with tab5:
        st.markdown("### Datos Históricos de Subgrupos")
        import pandas as pd
        rows_tbl = []
        n_sub = int(p["n_subgroup"])
        for i, sg in enumerate(p["subgroups"]):
            row = {"Subgrupo": i + 1}
            for j, v in enumerate(sg):
                row[f"Vol{j+1}"] = v
            row["X̄i (mm)"] = round(spc["means"][i], 2)
            row["Ri (mm)"]  = round(spc["ranges"][i], 2)
            row["Estado"]    = "⚠️ FUERA" if (spc["means"][i] > spc["UCL_X"] or
                                              spc["means"][i] < spc["LCL_X"]) else "✅ Control"
            rows_tbl.append(row)
        st.dataframe(pd.DataFrame(rows_tbl), use_container_width=True)

        st.markdown("### Estadísticas Monte Carlo")
        mc_tbl = []
        for lbl, st_d, unit in [
            ("X50 (mm)", mc["X50"], "mm"),
            ("PPV (mm/s)", mc["PPV"], "mm/s"),
            ("n uniformidad", mc["n_u"], "—"),
        ]:
            mc_tbl.append({
                "Variable": lbl, "Media": f"{st_d['mean']:.3f}",
                "σ": f"{st_d['std']:.3f}", "P5": f"{st_d['P5']:.3f}",
                "P50": f"{st_d['P50']:.3f}", "P90": f"{st_d['P90']:.3f}",
                "P95": f"{st_d['P95']:.3f}", "Unidad": unit,
            })
        st.dataframe(pd.DataFrame(mc_tbl), use_container_width=True)

    # ── Exportar Excel ────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("## 📥 Exportar Reporte")
    with st.spinner("Generando archivo Excel..."):
        xlsx_bytes = _build_excel(p, kuz, ppv_res, spc, cap, mc)
    st.download_button(
        label="⬇️  DESCARGAR REPORTE EXCEL (.xlsx)",
        data=xlsx_bytes,
        file_name="SPC_Voladura_Giovany_Valencia.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    # ── Footer ────────────────────────────────────────────────────────────
    st.markdown("""
<div style="
    margin-top:2rem;
    padding:1rem;
    border-top:1px solid #2D3550;
    text-align:center;
    font-family:'Share Tech Mono',monospace;
    font-size:0.72rem;
    color:#2D3550;
    letter-spacing:0.5px;">
  Kuz-Ram (Cunningham, 1987) &nbsp;·&nbsp; Sadovsky (1959) &nbsp;·&nbsp;
  SPC (Shewhart, 1931) &nbsp;·&nbsp; ISO 22514-2 &nbsp;·&nbsp;
  Montgomery (2013) &nbsp;·&nbsp; Giovany Valencia Huanca · UNA Puno · 2026
</div>
""", unsafe_allow_html=True)


if __name__ == "__main__":
    main()
